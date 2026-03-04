from __future__ import annotations

import io
import json
import os
import threading
import uuid
from datetime import datetime
from pathlib import Path
from statistics import mean

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "data.json"
DATABASE_URL = (os.getenv("DATABASE_URL") or "").strip()

WEIGHTS = {"C": 1.0, "I": 0.5, "S": 0.0}

app = Flask(__name__)
lock = threading.Lock()


def normalize_database_url(raw_url: str) -> str:
    if raw_url.startswith("postgres://"):
        return raw_url.replace("postgres://", "postgresql+psycopg://", 1)
    if raw_url.startswith("postgresql://") and "+" not in raw_url.split("://", 1)[0]:
        return raw_url.replace("postgresql://", "postgresql+psycopg://", 1)
    return raw_url


engine: Engine | None = None
if DATABASE_URL:
    engine = create_engine(normalize_database_url(DATABASE_URL), pool_pre_ping=True, future=True)


def default_state() -> dict:
    return {"courses": {}}


def init_database_if_needed() -> None:
    if not engine:
        return
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS app_state (
                    id INTEGER PRIMARY KEY,
                    data TEXT NOT NULL,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
        )
        row = conn.execute(text("SELECT id FROM app_state WHERE id = 1")).first()
        if not row:
            conn.execute(
                text("INSERT INTO app_state (id, data) VALUES (1, :data)"),
                {"data": json.dumps(default_state(), ensure_ascii=False)},
            )


init_database_if_needed()


def load_state() -> dict:
    if engine:
        try:
            with engine.begin() as conn:
                row = conn.execute(text("SELECT data FROM app_state WHERE id = 1")).first()
                if not row or not row[0]:
                    return default_state()
                return json.loads(row[0])
        except Exception:
            return default_state()

    if not DATA_FILE.exists():
        return default_state()
    try:
        return json.loads(DATA_FILE.read_text(encoding="utf-8"))
    except Exception:
        return default_state()


def save_state(state: dict) -> None:
    if engine:
        with engine.begin() as conn:
            payload = json.dumps(state, ensure_ascii=False)
            conn.execute(
                text(
                    """
                    INSERT INTO app_state (id, data, updated_at)
                    VALUES (1, :data, CURRENT_TIMESTAMP)
                    ON CONFLICT (id)
                    DO UPDATE SET data = EXCLUDED.data, updated_at = CURRENT_TIMESTAMP
                    """
                ),
                {"data": payload},
            )
        return

    DATA_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def get_state() -> dict:
    with lock:
        return load_state()


def update_state(mutator):
    with lock:
        state = load_state()
        result = mutator(state)
        save_state(state)
        return result


def calc_grade(marks: list[str], num_classes: int, exigency: float = 0.60) -> float | None:
    if not marks or num_classes <= 0:
        return None
    absence_count = sum(1 for m in marks if m == "A")
    effective_classes = max(0, num_classes - absence_count)
    if effective_classes <= 0:
        return None
    score = sum(0.0 if m == "A" else WEIGHTS.get(m, 0.0) for m in marks)
    max_score = float(effective_classes)
    passing_score = max_score * exigency
    if max_score == 0:
        return 7.0
    if score >= passing_score:
        grade = 4.0 + 3.0 * ((score - passing_score) / max(1e-9, (max_score - passing_score)))
    else:
        grade = 1.0 + 3.0 * (score / max(1e-9, passing_score))
    return max(1.0, min(7.0, grade))


def student_period_grade(course: dict, period: dict, student_name: str) -> float | None:
    classes = period.get("classes", [])
    if not classes:
        return None
    marks = period.get("marks", {}).get(student_name, [])
    exig = float(period.get("exigency", 60)) / 100.0
    return calc_grade(marks, len(classes), exig)


def student_average(course: dict, student_name: str) -> float | None:
    grades = []
    for period in course.get("periods", []):
        g = student_period_grade(course, period, student_name)
        if g is not None:
            grades.append(g)
    return mean(grades) if grades else None


def public_state(state: dict) -> dict:
    return state


def find_course(state: dict, course_name: str) -> dict:
    return state["courses"].get(course_name)


def find_period(course: dict, period_id: str) -> dict | None:
    for p in course.get("periods", []):
        if p.get("id") == period_id:
            return p
    return None


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/state", methods=["GET"])
def api_state():
    return jsonify(public_state(get_state()))


@app.route("/api/course", methods=["POST"])
def api_create_course():
    payload = request.get_json(force=True)
    name = (payload.get("name") or "").strip().upper()
    if not name:
        return jsonify({"error": "Nombre de curso requerido"}), 400

    def mut(state):
        if name in state["courses"]:
            return {"error": "El curso ya existe"}, 400
        state["courses"][name] = {"students": [], "periods": []}
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/course/<course_name>", methods=["DELETE"])
def api_delete_course(course_name: str):
    def mut(state):
        if course_name in state["courses"]:
            del state["courses"][course_name]
        return {"ok": True}

    return jsonify(update_state(mut))


@app.route("/api/course/<course_name>/student", methods=["POST"])
def api_add_student(course_name: str):
    payload = request.get_json(force=True)
    student_name = (payload.get("name") or "").strip()
    if not student_name:
        return jsonify({"error": "Nombre de estudiante requerido"}), 400

    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        if any(s.get("name") == student_name for s in course["students"]):
            return {"error": "Estudiante ya existe"}, 400
        course["students"].append({"name": student_name, "active": True})
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/course/<course_name>/student/<student_name>", methods=["DELETE"])
def api_remove_student(course_name: str, student_name: str):
    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        course["students"] = [s for s in course["students"] if s.get("name") != student_name]
        for p in course.get("periods", []):
            p.get("marks", {}).pop(student_name, None)
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/course/<course_name>/period", methods=["POST"])
def api_add_period(course_name: str):
    payload = request.get_json(force=True)
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Nombre de periodo requerido"}), 400

    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        course["periods"].append(
            {
                "id": f"p_{uuid.uuid4().hex[:8]}",
                "name": name,
                "exigency": 60,
                "classes": [],
                "marks": {},
            }
        )
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/course/<course_name>/period/<period_id>", methods=["PUT"])
def api_update_period(course_name: str, period_id: str):
    payload = request.get_json(force=True)

    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        period = find_period(course, period_id)
        if not period:
            return {"error": "Periodo no encontrado"}, 404
        if "name" in payload:
            period["name"] = (payload.get("name") or period["name"]).strip() or period["name"]
        if "exigency" in payload:
            ex = int(payload.get("exigency") or 60)
            period["exigency"] = max(1, min(100, ex))
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/course/<course_name>/period/<period_id>", methods=["DELETE"])
def api_delete_period(course_name: str, period_id: str):
    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        course["periods"] = [p for p in course["periods"] if p.get("id") != period_id]
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/course/<course_name>/period/<period_id>/class", methods=["POST"])
def api_add_class(course_name: str, period_id: str):
    payload = request.get_json(force=True)
    label = (payload.get("label") or "").strip() or "Clase"

    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        period = find_period(course, period_id)
        if not period:
            return {"error": "Periodo no encontrado"}, 404
        period["classes"].append({"label": label})
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/course/<course_name>/period/<period_id>/class/<int:class_idx>", methods=["PUT"])
def api_update_class(course_name: str, period_id: str, class_idx: int):
    payload = request.get_json(force=True)

    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        period = find_period(course, period_id)
        if not period:
            return {"error": "Periodo no encontrado"}, 404
        if class_idx < 0 or class_idx >= len(period["classes"]):
            return {"error": "Clase fuera de rango"}, 400
        period["classes"][class_idx]["label"] = (payload.get("label") or "").strip()
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/course/<course_name>/period/<period_id>/class/<int:class_idx>", methods=["DELETE"])
def api_delete_class(course_name: str, period_id: str, class_idx: int):
    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        period = find_period(course, period_id)
        if not period:
            return {"error": "Periodo no encontrado"}, 404
        if class_idx < 0 or class_idx >= len(period["classes"]):
            return {"error": "Clase fuera de rango"}, 400
        period["classes"].pop(class_idx)
        for student_name, marks in period.get("marks", {}).items():
            if class_idx < len(marks):
                marks.pop(class_idx)
            period["marks"][student_name] = marks
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/course/<course_name>/period/<period_id>/mark", methods=["POST"])
def api_set_mark(course_name: str, period_id: str):
    payload = request.get_json(force=True)
    student_name = payload.get("student")
    class_idx = int(payload.get("class_idx"))
    mark = payload.get("mark", "")
    if mark not in {"", "C", "I", "S", "A"}:
        return jsonify({"error": "Marca inválida"}), 400

    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        period = find_period(course, period_id)
        if not period:
            return {"error": "Periodo no encontrado"}, 404
        if class_idx < 0 or class_idx >= len(period.get("classes", [])):
            return {"error": "Clase fuera de rango"}, 400

        marks = period.setdefault("marks", {}).setdefault(student_name, [])
        while len(marks) <= class_idx:
            marks.append("")
        marks[class_idx] = mark
        return {"ok": True}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


@app.route("/api/export/json", methods=["GET"])
def api_export_json():
    state = get_state()
    payload = json.dumps(state, ensure_ascii=False, indent=2).encode("utf-8")
    return send_file(
        io.BytesIO(payload),
        as_attachment=True,
        download_name=f"respaldo_nota_proceso_{datetime.now().date()}.json",
        mimetype="application/json",
    )


@app.route("/api/import/json", methods=["POST"])
def api_import_json():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Archivo requerido"}), 400
    try:
        state = json.loads(file.read().decode("utf-8"))
    except Exception:
        return jsonify({"error": "JSON inválido"}), 400

    def mut(_):
        return state

    with lock:
        save_state(state)
    return jsonify({"ok": True})


@app.route("/api/course/<course_name>/upload-excel", methods=["POST"])
def api_upload_excel(course_name: str):
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "Archivo requerido"}), 400

    wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
    ws = wb[wb.sheetnames[0]]
    names = []
    for row in ws.iter_rows(min_row=1, max_col=1):
        value = row[0].value
        if value is None:
            continue
        n = str(value).strip()
        if n:
            names.append(n)

    def mut(state):
        course = find_course(state, course_name)
        if not course:
            return {"error": "Curso no encontrado"}, 404
        existing = {s.get("name") for s in course.get("students", [])}
        added = 0
        for n in names:
            if n not in existing:
                course["students"].append({"name": n, "active": True})
                existing.add(n)
                added += 1
        return {"ok": True, "added": added}

    result = update_state(mut)
    if isinstance(result, tuple):
        return jsonify(result[0]), result[1]
    return jsonify(result)


def _build_individual_pdf(course_name: str, course: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = []

    students = sorted([s["name"] for s in course.get("students", []) if s.get("active", True)])
    for idx, student_name in enumerate(students):
        if idx > 0:
            story.append(Spacer(1, 24))
        story.append(Paragraph(f"<b>Informe Individual - {course_name}</b>", styles["Title"]))
        story.append(Paragraph(f"Estudiante: <b>{student_name}</b>", styles["Normal"]))
        avg = student_average(course, student_name)
        story.append(Paragraph(f"Promedio histórico: <b>{avg:.1f}</b>" if avg else "Promedio histórico: -", styles["Normal"]))
        story.append(Spacer(1, 8))

        for period in course.get("periods", []):
            classes = period.get("classes", [])
            if not classes:
                continue
            marks = period.get("marks", {}).get(student_name, [])
            grade = student_period_grade(course, period, student_name)
            story.append(Paragraph(f"<b>{period.get('name')}</b>", styles["Heading4"]))
            table_data = [["Clase", "Etiqueta", "Marca", "Puntaje"]]
            for i, cls in enumerate(classes):
                mark = marks[i] if i < len(marks) else ""
                if mark == "A":
                    pts = "N/A"
                else:
                    pts = WEIGHTS.get(mark, "-") if mark else "-"
                table_data.append([f"C{i+1}", cls.get("label", ""), mark or "-", str(pts)])
            table_data.append(["", "", "NOTA", f"{grade:.1f}" if grade is not None else "-"])

            t = Table(table_data, colWidths=[45, 250, 60, 60])
            t.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ]
                )
            )
            story.append(t)
            story.append(Spacer(1, 10))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def _build_group_pdf(course_name: str, course: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    story = [Paragraph(f"<b>Informe Grupal - {course_name}</b>", styles["Title"]), Spacer(1, 10)]

    students = sorted([s["name"] for s in course.get("students", []) if s.get("active", True)])
    periods = course.get("periods", [])
    headers = ["Estudiante"] + [p.get("name", "") for p in periods] + ["Promedio"]
    data = [headers]
    for s in students:
        row = [s]
        for p in periods:
            g = student_period_grade(course, p, s)
            row.append(f"{g:.1f}" if g is not None else "-")
        avg = student_average(course, s)
        row.append(f"{avg:.1f}" if avg is not None else "-")
        data.append(row)

    col_widths = [170] + [80] * len(periods) + [80]
    table = Table(data, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


@app.route("/api/course/<course_name>/report/individual.pdf", methods=["GET"])
def api_pdf_individual(course_name: str):
    state = get_state()
    course = find_course(state, course_name)
    if not course:
        return jsonify({"error": "Curso no encontrado"}), 404

    content = _build_individual_pdf(course_name, course)
    return send_file(
        io.BytesIO(content),
        as_attachment=True,
        download_name=f"nota_proceso_individual_{course_name}.pdf",
        mimetype="application/pdf",
    )


@app.route("/api/course/<course_name>/report/group.pdf", methods=["GET"])
def api_pdf_group(course_name: str):
    state = get_state()
    course = find_course(state, course_name)
    if not course:
        return jsonify({"error": "Curso no encontrado"}), 404

    content = _build_group_pdf(course_name, course)
    return send_file(
        io.BytesIO(content),
        as_attachment=True,
        download_name=f"nota_proceso_grupal_{course_name}.pdf",
        mimetype="application/pdf",
    )


if __name__ == "__main__":
    if not engine and not DATA_FILE.exists():
        save_state(default_state())
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
